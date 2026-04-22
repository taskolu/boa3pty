import csv
from decimal import Decimal
from datetime import datetime
from typing import Optional
from pathlib import Path
from src.core.models import GPGPayment

_XLS_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",  # 2026-03-18 00:00:00
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%d %b %Y",
    "%d-%b-%Y",
    "%Y/%m/%d",
]


def _parse_date(raw: str, fmt: str):
    """Try configured format first, then all known formats."""
    raw = raw.strip()
    # Strip time suffix if present (e.g. "2026-03-18 00:00:00" → "2026-03-18")
    if " " in raw and fmt == "%Y-%m-%d":
        raw = raw.split(" ")[0]
    try:
        return datetime.strptime(raw, fmt).date()
    except ValueError:
        pass
    for f in _XLS_DATE_FORMATS:
        try:
            return datetime.strptime(raw, f).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse date: {raw!r}")


def _read_rows(file_path: str, delimiter: str = ",") -> tuple[list[dict], list[str]]:
    """Read a CSV, XLS, or XLSX file into a list of row dicts.

    Returns (rows, headers).
    """
    ext = Path(file_path).suffix.lower()

    if ext in (".xls",):
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        headers = [str(c).strip() for c in ws.row_values(0)]
        rows = []
        for r in range(1, ws.nrows):
            vals = ws.row_values(r)
            rows.append({headers[i]: str(vals[i]).strip() for i in range(len(headers))})
        return rows, headers

    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        rows = []
        for row in rows_iter:
            rows.append({headers[i]: str(row[i]).strip() if row[i] is not None else ""
                         for i in range(len(headers))})
        wb.close()
        return rows, headers

    # CSV (comma, semicolon, or tab)
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        rows = list(reader)
    return rows, headers


def parse_gpg_file(
    file_path: str,
    column_mapping: dict,
    date_format: str,
    bank_code_column: str = "inventory_code",
    delimiter: str = ",",
) -> tuple[list[GPGPayment], Optional[str]]:
    """Parse a GPG payment report (CSV, XLS, or XLSX) into GPGPayment records.

    Returns (list of payments, detected bank code).
    """
    rows, _ = _read_rows(file_path, delimiter)
    records = []
    detected_bank_code = None

    for row in rows:
        # Detect bank code from first row
        if detected_bank_code is None and bank_code_column in row:
            val = row[bank_code_column].strip()
            if val:
                detected_bank_code = val

        # Status: prefer error_code if present, else payment_status
        raw_status = ""
        for sc_col in [column_mapping.get("status_code", ""), "error_code", "payment_status"]:
            if sc_col and sc_col in row:
                raw_status = row[sc_col].strip()
                if raw_status:
                    break

        # Also include error_description as status_message if available
        status_code = None
        status_message = None
        if raw_status:
            parts = raw_status.split(",", 1)
            status_code = parts[0].strip()
            status_message = parts[1].strip() if len(parts) > 1 else None

        if not status_message:
            desc_col = "error_description"
            if desc_col in row and row[desc_col].strip():
                status_message = row[desc_col].strip()

        # Skip rows with no amount (summary/blank rows in Excel)
        raw_amount = row.get(column_mapping.get("buy_amount", ""), "").strip().replace(",", "")
        if not raw_amount:
            continue

        payment = GPGPayment(
            payment_id=row.get(column_mapping.get("payment_id", ""), "").strip()
                       or row.get(column_mapping.get("confirmation_number", ""), "").strip(),
            confirmation_number=row.get(column_mapping["confirmation_number"], "").strip(),
            buy_currency=row.get(column_mapping["buy_currency"], "").strip(),
            buy_amount=Decimal(raw_amount),
            value_date=_parse_date(
                row.get(column_mapping["value_date"], "").strip(), date_format
            ),
            status_code=status_code,
            status_message=status_message,
            counterparty=row.get(bank_code_column, "").strip(),
            raw_row=dict(row),
        )
        records.append(payment)

    return records, detected_bank_code


# Keep old name as alias so existing tests still work
def parse_gpg_csv(
    file_path: str,
    column_mapping: dict,
    date_format: str,
    bank_code_column: str = "Bank Code",
    delimiter: str = ",",
) -> tuple[list[GPGPayment], Optional[str]]:
    return parse_gpg_file(file_path, column_mapping, date_format, bank_code_column, delimiter)
