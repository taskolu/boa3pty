from datetime import date
from decimal import Decimal
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from src.core.models import MatchResult


def generate_payment_breakdown(
    results: list[MatchResult],
    output_path: str,
    report_date: date,
):
    """Export a payment breakdown grouped by currency to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Breakdown"

    ws.append(["Payment Breakdown Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Date:", report_date.isoformat()])
    ws.append(["Total Records:", len(results)])
    ws.append([])

    by_ccy: dict[str, list[MatchResult]] = defaultdict(list)
    for r in results:
        if r.gpg_record:
            by_ccy[r.gpg_record.buy_currency].append(r)

    headers = [
        "Confirmation#", "Status", "Currency", "Amount",
        "Value Date (GPG)", "Value Date (WS)", "WS Deal #", "WS Ext Deal #"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for ccy, records in sorted(by_ccy.items()):
        ws.append([f"Currency: {ccy}"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        row_num = ws.max_row + 1
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font

        total = Decimal("0")
        for r in records:
            gpg = r.gpg_record
            ws_entry = r.ws_record
            ws.append([
                gpg.confirmation_number,
                r.status.value,
                gpg.buy_currency,
                float(gpg.buy_amount),
                gpg.value_date.isoformat(),
                ws_entry.value_date.isoformat() if ws_entry else "",
                ws_entry.wallstreet_ref if ws_entry else "",
                ws_entry.external_ref if ws_entry else "",
            ])
            total += gpg.buy_amount

        ws.append(["", "", "TOTAL:", float(total)])
        ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
        ws.append([])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    wb.save(output_path)
