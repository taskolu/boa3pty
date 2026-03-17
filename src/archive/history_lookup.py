from datetime import date, timedelta
from pathlib import Path
from openpyxl import load_workbook


def lookup_flagged_records(
    archive_path: str,
    counterparty: str,
    lookback_days: int = 5,
    reference_date: date | None = None,
) -> list[dict]:
    """Scan recent archive files for flagged (DT06 / unmatched) records.

    Returns a list of dicts with confirmation_number, status, source_file, etc.
    Used by the matcher to resolve previously flagged DT06 entries.
    """
    ref = reference_date or date.today()
    base = Path(archive_path)
    flags: list[dict] = []

    for i in range(1, lookback_days + 1):
        dt = ref - timedelta(days=i)
        filepath = base / f"{dt.isoformat()}_{counterparty}.xlsx"
        if not filepath.exists():
            continue

        wb = load_workbook(str(filepath), read_only=True)
        if "Flagged" not in wb.sheetnames:
            wb.close()
            continue

        ws = wb["Flagged"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            wb.close()
            continue

        headers = rows[0]
        for row in rows[1:]:
            record = dict(zip(headers, row))
            flags.append({
                "confirmation_number": record.get("Confirmation#", ""),
                "status": record.get("Status", ""),
                "currency": record.get("Currency", ""),
                "amount": record.get("Amount", ""),
                "value_date": record.get("ValueDate", ""),
                "source_file": str(filepath),
            })

        wb.close()

    return flags
