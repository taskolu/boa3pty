from __future__ import annotations
import os
import shutil
import tempfile
import time
from datetime import date
from pathlib import Path
from typing import Optional
from openpyxl import Workbook, load_workbook
from src.core.models import MatchResult, MatchStatus


class ArchiveManager:
    def __init__(self, archive_path: str):
        self._path = Path(archive_path)
        self._path.mkdir(parents=True, exist_ok=True)

    def _filename(self, dt: date, counterparty: str) -> Path:
        return self._path / f"{dt.isoformat()}_{counterparty}.xlsx"

    def save_daily(
        self, dt: date, counterparty: str, results: list[MatchResult]
    ):
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        counts: dict[str, int] = {}
        for r in results:
            key = r.status.value
            counts[key] = counts.get(key, 0) + 1
        ws_summary.append(["Date", dt.isoformat()])
        ws_summary.append(["Counterparty", counterparty])
        ws_summary.append(["Total Records", len(results)])
        for status, count in counts.items():
            ws_summary.append([status, count])

        # Results sheet
        ws_results = wb.create_sheet("Results")
        ws_results.append([
            "Status", "Confirmation#", "GPG_Currency", "GPG_Amount",
            "GPG_ValueDate", "GPG_StatusCode",
            "WS_RecCcy", "WS_RecAmount", "WS_ValueDate", "WS_Ref",
            "WS_PayCcy", "WS_PayAmount", "WS_Rate",
            "ClientAccount", "ArrivalDate",
            "Discrepancies", "Resolution_Source"
        ])
        for r in results:
            gpg = r.gpg_record
            ws  = r.ws_record
            raw = gpg.raw_row if gpg else {}
            arrival = raw.get("Arrival_date_in_UTC", raw.get("arrival_date", ""))
            arrival = arrival.split(" ")[0] if arrival else ""
            client  = raw.get("client_account_number", "")
            ws_results.append([
                r.status.value,
                (gpg.confirmation_number if gpg
                 else ws.external_ref if ws else ""),
                gpg.buy_currency if gpg else "",
                str(gpg.buy_amount) if gpg else "",
                gpg.value_date.isoformat() if gpg else "",
                gpg.status_code or "" if gpg else "",
                ws.rec_ccy if ws else "",
                str(ws.rec_amount) if ws else "",
                ws.value_date.isoformat() if ws else "",
                ws.wallstreet_ref if ws else "",
                ws.pay_ccy if ws else "",
                str(ws.pay_amount) if ws else "",
                str(ws.rate) if ws else "",
                client,
                arrival,
                "; ".join(r.discrepancies),
                r.resolution_source or "",
            ])

        # Flagged sheet (used for DT06 lookback)
        ws_flagged = wb.create_sheet("Flagged")
        ws_flagged.append([
            "Confirmation#", "Status", "Currency", "Amount",
            "ValueDate", "StatusCode", "StatusMessage"
        ])
        for r in results:
            if r.status in (MatchStatus.FLAGGED_DT06, MatchStatus.UNMATCHED_GPG):
                gpg = r.gpg_record
                if gpg:
                    ws_flagged.append([
                        gpg.confirmation_number,
                        r.status.value,
                        gpg.buy_currency,
                        str(gpg.buy_amount),
                        gpg.value_date.isoformat(),
                        gpg.status_code or "",
                        gpg.status_message or "",
                    ])

        filepath = self._filename(dt, counterparty)
        # Write to system temp dir (outside OneDrive) so OneDrive cannot grab
        # the temp file mid-write, then copy to the final OneDrive destination
        # with retries to handle transient sync locks (WinError 32).
        tmp = tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=tempfile.gettempdir(), delete=False
        )
        tmp.close()
        try:
            wb.save(tmp.name)
            last_err = None
            for _ in range(5):
                try:
                    shutil.copy2(tmp.name, str(filepath))
                    last_err = None
                    break
                except OSError as e:
                    last_err = e
                    time.sleep(1.0)
            if last_err:
                raise last_err
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    def load_daily(self, dt: date, counterparty: str) -> Optional[dict]:
        filepath = self._filename(dt, counterparty)
        if not filepath.exists():
            return None

        wb = load_workbook(str(filepath), read_only=True)
        ws_summary = wb["Summary"]
        summary: dict = {}
        for row in ws_summary.iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                summary[row[0]] = row[1]
        wb.close()
        return {"summary": summary, "file": str(filepath)}

    def load_results_sheet(self, filepath: str) -> list[dict]:
        """Read the Results sheet from an archive file and return list of row dicts.

        Copies the file to a temp location first so OneDrive sync locks don't
        cause PermissionError on the shared OneDrive folder.
        Retries up to 3 times with a short delay to handle transient locks.
        """
        last_err = None
        for attempt in range(3):
            try:
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp.close()
                shutil.copy2(filepath, tmp.name)
                try:
                    wb = load_workbook(tmp.name, read_only=True)
                    if "Results" not in wb.sheetnames:
                        wb.close()
                        return []
                    ws = wb["Results"]
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                finally:
                    os.unlink(tmp.name)
                if len(rows) < 2:
                    return []
                headers = [str(h) if h is not None else "" for h in rows[0]]
                return [dict(zip(headers, row)) for row in rows[1:]]
            except PermissionError as e:
                last_err = e
                time.sleep(1.5)
        raise last_err

    def list_archives(self) -> list[dict]:
        archives = []
        for f in sorted(self._path.glob("*.xlsx"), reverse=True):
            parts = f.stem.split("_", 1)
            if len(parts) == 2:
                archives.append({
                    "date": parts[0],
                    "counterparty": parts[1],
                    "file": str(f),
                })
        return archives
