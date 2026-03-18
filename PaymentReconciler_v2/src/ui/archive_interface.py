from __future__ import annotations
import time

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QHeaderView,
    QAbstractItemView, QFileDialog, QMessageBox, QDateEdit
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QFont, QBrush

from qfluentwidgets import (
    PushButton, SubtitleLabel, BodyLabel,
    TableWidget, ComboBox
)

from src.archive.archive_manager import ArchiveManager
from src.core.app_dir import resolve_archive_path
from src.export.report_generator import generate_payment_breakdown


def _fmt_rate(v) -> str:
    try:
        from decimal import Decimal
        return format(Decimal(str(v)).normalize(), 'f')
    except Exception:
        return str(v) if v else ""

# ── Colours (mirrors reconcile) ─────────────────────────────────────────────
_WS_HDR_BG  = QColor("#1a2a3a")
_KEY_HDR_BG = QColor("#3a2a00")
_GPG_HDR_BG = QColor("#1a3a1a")

_STATUS_BG = {
    "matched":               QColor("#152815"),
    "unmatched_gpg":         QColor("#301010"),
    "unmatched_ws":          QColor("#2a1800"),
    "flagged_dt06":          QColor("#181828"),
    "resolved_from_archive": QColor("#152015"),
    "amount_mismatch":       QColor("#2a1200"),
    "currency_mismatch":     QColor("#2a1200"),
    "value_date_mismatch":   QColor("#1a1a2e"),
}
_STATUS_FG = {
    "matched":               QColor("#4caf50"),
    "unmatched_gpg":         QColor("#ef5350"),
    "unmatched_ws":          QColor("#ff9800"),
    "flagged_dt06":          QColor("#9fa8da"),
    "resolved_from_archive": QColor("#81c784"),
    "amount_mismatch":       QColor("#ff7043"),
    "currency_mismatch":     QColor("#ff7043"),
    "value_date_mismatch":   QColor("#ce93d8"),
}
_STATUS_LABEL = {
    "matched":               "Matched",
    "unmatched_gpg":         "Missing in WS",
    "unmatched_ws":          "Extra in WS",
    "flagged_dt06":          "DT06",
    "resolved_from_archive": "Resolved",
    "amount_mismatch":       "Amt Mismatch",
    "currency_mismatch":     "Ccy Mismatch",
    "value_date_mismatch":   "Date Mismatch",
}

_HEADERS = [
    "Status",
    "Value Date",
    "Pay Ccy",
    "Pay Amount",
    "Rate",
    "Buy Ccy",
    "Buy Amount",
    "WS Deal #",
    "Conf # / Ext Deal #",
    "GPG Status",
    "GPG Value Date",
    "GPG Amount",
    "Currency",
    "Client Account",
    "Arrival Date",
    "Notes",
]
_CONF_COL = 8
_WS_COLS  = [1, 2, 3, 4, 5, 6, 7]
_GPG_COLS = [9, 10, 11, 12, 13, 14, 15]


class ArchiveInterface(QWidget):
    def __init__(self, config_manager, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("archiveInterface")
        self.config = config_manager
        self._archive_list: list[dict] = []
        self._current_rows: list[dict] = []
        self._last_refresh: float = 0.0
        self._loaded_filepath: str = ""
        self._init_ui()
        self.refresh()

    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 24, 24, 24)
        root.setSpacing(12)

        root.addWidget(SubtitleLabel("Archive"))

        # ── Controls ──────────────────────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.addWidget(BodyLabel("Value Date:"))
        self.dt_picker = QDateEdit(self)
        self.dt_picker.setDisplayFormat("dd MMM yyyy")
        self.dt_picker.setCalendarPopup(True)
        self.dt_picker.setDate(QDate.currentDate())
        self.dt_picker.dateChanged.connect(self._on_selection_changed)
        ctrl.addWidget(self.dt_picker)

        ctrl.addSpacing(12)
        ctrl.addWidget(BodyLabel("Counterparty:"))
        self.cmb_cp = ComboBox(self)
        self.cmb_cp.setMinimumWidth(160)
        self.cmb_cp.currentIndexChanged.connect(self._load_selected)
        ctrl.addWidget(self.cmb_cp)

        self.btn_refresh = PushButton("Refresh", self)
        self.btn_refresh.clicked.connect(self.refresh)
        ctrl.addWidget(self.btn_refresh)

        ctrl.addStretch()

        self.lbl_count = BodyLabel("")
        ctrl.addWidget(self.lbl_count)

        self.btn_export = PushButton("Export to Excel…", self)
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._export)
        ctrl.addWidget(self.btn_export)

        root.addLayout(ctrl)

        # ── Section hints ─────────────────────────────────────────────
        hint = QHBoxLayout()
        ws_lbl = BodyLabel("◀  WallStreet")
        ws_lbl.setStyleSheet("color: #5b9bd5; font-weight: bold; font-size: 10px;")
        hint.addWidget(ws_lbl)
        hint.addStretch()
        key_lbl = BodyLabel("MATCHING KEY")
        key_lbl.setStyleSheet("color: #ffc107; font-weight: bold; font-size: 10px;")
        hint.addWidget(key_lbl)
        hint.addStretch()
        gpg_lbl = BodyLabel("GPG  ▶")
        gpg_lbl.setStyleSheet("color: #66bb6a; font-weight: bold; font-size: 10px;")
        hint.addWidget(gpg_lbl)
        root.addLayout(hint)

        # ── Table ─────────────────────────────────────────────────────
        self.tbl = TableWidget(self)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setShowGrid(True)
        self.tbl.setWordWrap(False)
        self.tbl.setColumnCount(len(_HEADERS))
        self.tbl.setHorizontalHeaderLabels(_HEADERS)
        self._style_header()
        root.addWidget(self.tbl, 1)

    def _style_header(self):
        from PySide6.QtWidgets import QTableWidgetItem
        for col in range(len(_HEADERS)):
            item = QTableWidgetItem(_HEADERS[col])
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Segoe UI", 8, QFont.Bold))
            if col in _WS_COLS:
                item.setBackground(QBrush(_WS_HDR_BG))
                item.setForeground(QBrush(QColor("#5b9bd5")))
            elif col == _CONF_COL:
                item.setBackground(QBrush(_KEY_HDR_BG))
                item.setForeground(QBrush(QColor("#ffc107")))
            elif col in _GPG_COLS:
                item.setBackground(QBrush(_GPG_HDR_BG))
                item.setForeground(QBrush(QColor("#66bb6a")))
            self.tbl.setHorizontalHeaderItem(col, item)

    # ── Data loading ───────────────────────────────────────────────

    def showEvent(self, event):
        super().showEvent(event)
        if time.monotonic() - self._last_refresh >= 30.0:
            self.refresh()

    def reload_config(self):
        self.refresh()

    def refresh(self):
        self._last_refresh = time.monotonic()
        self._loaded_filepath = ""
        try:
            archive_path = resolve_archive_path(self.config.archive_path)
            am = ArchiveManager(archive_path)
            self._archive_list = am.list_archives()
        except Exception:
            self._archive_list = []

        current = self._current_date_str()
        available_dates = [a["date"] for a in self._archive_list]
        if available_dates and current not in available_dates:
            newest = available_dates[0]
            try:
                parts = newest.split("-")
                self.dt_picker.blockSignals(True)
                self.dt_picker.setDate(QDate(int(parts[0]), int(parts[1]), int(parts[2])))
                self.dt_picker.blockSignals(False)
            except Exception:
                pass

        self._on_selection_changed()

    def _current_date_str(self) -> str:
        qd = self.dt_picker.date()
        return f"{qd.year():04d}-{qd.month():02d}-{qd.day():02d}"

    def _on_selection_changed(self):
        chosen_date = self._current_date_str()
        cps = [a["counterparty"] for a in self._archive_list if a["date"] == chosen_date]

        self.cmb_cp.blockSignals(True)
        self.cmb_cp.clear()
        self.cmb_cp.addItems(cps)
        self.cmb_cp.blockSignals(False)

        self._load_selected()

    def _load_selected(self):
        chosen_date = self._current_date_str()
        chosen_cp   = self.cmb_cp.currentText()
        if not chosen_date or not chosen_cp:
            self._current_rows = []
            self._populate_table([])
            return

        filepath = None
        for a in self._archive_list:
            if a["date"] == chosen_date and a["counterparty"] == chosen_cp:
                filepath = a["file"]
                break
        if not filepath:
            return

        if filepath == self._loaded_filepath and self._current_rows:
            self._populate_table(self._current_rows)
            self.btn_export.setEnabled(True)
            return

        try:
            archive_path = resolve_archive_path(self.config.archive_path)
            am = ArchiveManager(archive_path)
            rows = am.load_results_sheet(filepath)
        except Exception as e:
            QMessageBox.warning(self, "Load Error", str(e))
            rows = []

        def _sort_key(r):
            sv = str(r.get("Status", "")).lower()
            is_matched = 0 if sv == "matched" else 1
            ccy = str(r.get("WS_RecCcy", "") or "")
            try:
                amt = float(str(r.get("WS_RecAmount", 0) or 0).replace(",", ""))
            except ValueError:
                amt = 0.0
            return (is_matched, ccy, amt)

        rows = sorted(rows, key=_sort_key)
        rows = self._annotate_dt06_resolution(rows, chosen_date, am)

        self._current_rows = rows
        self._loaded_filepath = filepath
        self._populate_table(self._current_rows)
        self.btn_export.setEnabled(bool(self._current_rows))

    def _annotate_dt06_resolution(self, rows: list[dict], current_date: str,
                                   am: ArchiveManager) -> list[dict]:
        dt06_confs = {
            str(r.get("Confirmation#", "") or "")
            for r in rows
            if str(r.get("Status", "")).lower() == "flagged_dt06"
        }
        if not dt06_confs:
            return rows

        later_archives = sorted(
            [a for a in self._archive_list if a["date"] > current_date],
            key=lambda a: a["date"]
        )

        resolution_map: dict[str, str] = {}
        for arch in later_archives:
            if not dt06_confs - set(resolution_map.keys()):
                break
            try:
                future_rows = am.load_results_sheet(arch["file"])
            except Exception:
                continue
            date_label = arch["date"]
            try:
                parts = date_label.split("-")
                from datetime import date as _date
                d = _date(int(parts[0]), int(parts[1]), int(parts[2]))
                date_label = d.strftime("%d-%b")
            except Exception:
                pass
            for fr in future_rows:
                conf = str(fr.get("Confirmation#", "") or "")
                if conf in dt06_confs and conf not in resolution_map:
                    sv   = str(fr.get("Status", "")).lower()
                    disc = str(fr.get("Discrepancies", "") or "")
                    if sv == "matched" and "Bank amended value date" in disc:
                        resolution_map[conf] = f"→ ✓ Matched on {date_label} (bank amended value date)"
                    else:
                        label = _STATUS_LABEL.get(sv, sv)
                        resolution_map[conf] = f"→ {label} on {date_label}"

        if not resolution_map:
            return rows

        annotated = []
        for r in rows:
            sv = str(r.get("Status", "")).lower()
            if sv == "flagged_dt06":
                conf = str(r.get("Confirmation#", "") or "")
                res = resolution_map.get(conf, "→ Not resolved in later archives")
                r = dict(r)
                existing = str(r.get("Discrepancies", "") or "")
                r["Discrepancies"] = f"{existing}  {res}".strip() if existing else res
            annotated.append(r)
        return annotated

    def _populate_table(self, rows: list[dict]):
        from PySide6.QtWidgets import QTableWidgetItem
        self.tbl.setRowCount(len(rows))
        self.lbl_count.setText(f"{len(rows)} records" if rows else "")

        for ri, r in enumerate(rows):
            sv        = str(r.get("Status", "")).lower()
            row_bg    = _STATUS_BG.get(sv, QColor("#252525"))
            status_fg = _STATUS_FG.get(sv, QColor("white"))
            key_bg    = QColor("#2a1e00") if sv == "matched" else QColor("#1e1a00")
            conf_fg   = QColor("#4caf50") if sv == "matched" else QColor("#ffc107")

            conf   = str(r.get("Confirmation#", "") or "")
            vd_ws  = str(r.get("WS_ValueDate", "") or "")
            vd_gpg = str(r.get("GPG_ValueDate", "") or "")

            row_data = [
                (_STATUS_LABEL.get(sv, sv),                    row_bg, status_fg, True),
                (vd_ws,                                        row_bg, None,      False),
                (str(r.get("WS_PayCcy", "") or ""),            row_bg, None,      False),
                (str(r.get("WS_PayAmount", "") or ""),         row_bg, None,      False),
                (_fmt_rate(r.get("WS_Rate", "")),              row_bg, None,      False),
                (str(r.get("WS_RecCcy", "") or ""),            row_bg, None,      False),
                (str(r.get("WS_RecAmount", "") or ""),         row_bg, None,      False),
                (str(r.get("WS_Ref", "") or ""),               row_bg, None,      False),
                (conf,                                         key_bg, conf_fg,   True),
                (str(r.get("GPG_StatusCode", "") or ""),       row_bg, None,      False),
                (vd_gpg,                                       row_bg, None,      False),
                (str(r.get("GPG_Amount", "") or ""),           row_bg, None,      False),
                (str(r.get("GPG_Currency", "") or ""),         row_bg, None,      False),
                (str(r.get("ClientAccount", "") or ""),        row_bg, None,      False),
                (str(r.get("ArrivalDate", "") or ""),          row_bg, None,      False),
                (str(r.get("Discrepancies", "") or ""),        row_bg, None,      False),
            ]

            for ci, (val, bg, fg, bold) in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setBackground(QBrush(bg))
                if fg:
                    item.setForeground(QBrush(fg))
                if bold:
                    item.setFont(QFont("Segoe UI", 9, QFont.Bold))
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.tbl.setItem(ri, ci, item)

        self.tbl.resizeColumnsToContents()
        hh = self.tbl.horizontalHeader()
        for col in range(self.tbl.columnCount() - 1):
            if hh.sectionSize(col) > 220:
                hh.resizeSection(col, 220)

    def _export(self):
        if not self._current_rows:
            return
        chosen_date = self._current_date_str()
        chosen_cp   = self.cmb_cp.currentText()
        default_name = f"{chosen_date}_{chosen_cp}.xlsx" if chosen_date and chosen_cp else "archive.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Archive", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook as WB
            from openpyxl.styles import Font, PatternFill
            wb = WB()
            ws = wb.active
            ws.title = "Archive"

            hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            col_names = [
                "Status", "Value Date (WS)", "Pay Ccy", "Pay Amount", "Rate",
                "Buy Ccy", "Buy Amount", "WS Deal #", "Conf #",
                "GPG Status", "GPG Value Date", "GPG Amount", "Currency",
                "Client Account", "Arrival Date", "Notes"
            ]
            ws.append(col_names)
            for col_idx in range(1, len(col_names) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = hdr_fill
                cell.font = hdr_font

            for r in self._current_rows:
                sv = str(r.get("Status", ""))
                ws.append([
                    _STATUS_LABEL.get(sv.lower(), sv),
                    str(r.get("WS_ValueDate", "") or ""),
                    str(r.get("WS_PayCcy", "") or ""),
                    str(r.get("WS_PayAmount", "") or ""),
                    str(r.get("WS_Rate", "") or ""),
                    str(r.get("WS_RecCcy", "") or ""),
                    str(r.get("WS_RecAmount", "") or ""),
                    str(r.get("WS_Ref", "") or ""),
                    str(r.get("Confirmation#", "") or ""),
                    str(r.get("GPG_StatusCode", "") or ""),
                    str(r.get("GPG_ValueDate", "") or ""),
                    str(r.get("GPG_Amount", "") or ""),
                    str(r.get("GPG_Currency", "") or ""),
                    str(r.get("ClientAccount", "") or ""),
                    str(r.get("ArrivalDate", "") or ""),
                    str(r.get("Discrepancies", "") or ""),
                ])

            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            wb.save(path)
            QMessageBox.information(self, "Export", f"Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
