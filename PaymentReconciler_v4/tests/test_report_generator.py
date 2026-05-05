import sys
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.core.models import GPGPayment, MatchResult, MatchStatus, WSEntry
from src.export.report_generator import generate_payment_breakdown


def gpg(conf, amount="1000.00"):
    return GPGPayment(
        payment_id=conf,
        confirmation_number=conf,
        buy_currency="AZN",
        buy_amount=Decimal(amount),
        value_date=date(2026, 4, 30),
        status_code=None,
        status_message=None,
        counterparty="BOA3PTY",
    )


def ws(conf, rec_amount="4439.00", pay_amount="2667.35"):
    return WSEntry(
        value_date=date(2026, 4, 30),
        counterparty="BOA Exotic",
        pay_ccy="USD",
        pay_amount=Decimal(pay_amount),
        rec_ccy="AZN",
        rec_amount=Decimal(rec_amount),
        rate=Decimal("1.6642"),
        trader="",
        wallstreet_ref="W260428028669",
        external_ref=conf,
    )


class ReportGeneratorTests(unittest.TestCase):
    def test_gpg_only_missing_rows_do_not_affect_net_totals(self):
        results = [
            MatchResult(MatchStatus.UNMATCHED_GPG, gpg("MISSING", "1000.00"), None),
            MatchResult(MatchStatus.MATCHED, gpg("MATCHED", "4439.00"), ws("MATCHED")),
        ]

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report.xlsx"
            generate_payment_breakdown(results, str(output), date(2026, 4, 30))

            wb = load_workbook(output, data_only=True)
            sheet = wb["Payment Breakdown"]
            totals = {}
            for row in range(6, sheet.max_row + 1):
                ccy = sheet.cell(row=row, column=13).value
                total = sheet.cell(row=row, column=14).value
                if ccy:
                    totals[ccy] = Decimal(str(total))
            wb.close()

        self.assertEqual(totals["AZN"], Decimal("4439"))
        self.assertEqual(totals["USD"], Decimal("-2667.35"))


if __name__ == "__main__":
    unittest.main()
