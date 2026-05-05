import sys
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.core.matcher import reconcile
from src.core.models import MatchResult, MatchStatus, WSEntry
from src.core.parser_wallstreet import parse_wallstreet_paste
from src.export.report_generator import generate_payment_breakdown


def ws_blank_ext():
    return WSEntry(
        value_date=date(2026, 4, 30),
        counterparty="BOA Exotic",
        pay_ccy="USD",
        pay_amount=Decimal("2667.35"),
        rec_ccy="AZN",
        rec_amount=Decimal("4439.00"),
        rate=Decimal("1.6642"),
        trader="",
        wallstreet_ref="W260428028669",
        external_ref="",
    )


class BlankWallStreetExternalRefTests(unittest.TestCase):
    def test_parser_keeps_ws_row_with_blank_ext_deal(self):
        pasted = "\t".join([
            "Deal Type", "Value Date", "Customer", "Pay Ccy", "Pay Amount",
            "Rec Ccy", "Rec Amount", "Rate", "Trader", "Deal #", "Ext Deal #",
        ])
        pasted += "\n" + "\t".join([
            "FX", "30 Apr 2026", "BOA Exotic", "USD", "2,667.35",
            "AZN", "4,439.00", "1.6642", "", "W260428028669", "",
        ])

        entries, _ = parse_wallstreet_paste(pasted)

        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].external_ref, "")
        self.assertEqual(entries[0].wallstreet_ref, "W260428028669")

    def test_blank_ext_deal_reconciles_as_extra_in_ws(self):
        results = reconcile([], [ws_blank_ext()])

        self.assertEqual(results[0].status, MatchStatus.UNMATCHED_WS)
        self.assertEqual(results[0].ws_record.wallstreet_ref, "W260428028669")

    def test_export_uses_ws_deal_as_key_when_ext_deal_is_blank(self):
        results = [MatchResult(MatchStatus.UNMATCHED_WS, None, ws_blank_ext())]

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report.xlsx"
            generate_payment_breakdown(results, str(output), date(2026, 4, 30))

            wb = load_workbook(output, data_only=True)
            sheet = wb["Payment Breakdown"]
            conf_key = sheet.cell(row=6, column=1).value
            ws_deal = sheet.cell(row=6, column=10).value
            ws_ext = sheet.cell(row=6, column=11).value
            wb.close()

        self.assertEqual(conf_key, "W260428028669")
        self.assertEqual(ws_deal, "W260428028669")
        self.assertIsNone(ws_ext)


if __name__ == "__main__":
    unittest.main()
