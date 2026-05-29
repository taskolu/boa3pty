from datetime import date
from decimal import Decimal
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from src.core.models import MatchResult, MatchStatus


_HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FONT = Font(bold=True)
_NET_HDR_FILL = PatternFill(start_color="2E4A7A", end_color="2E4A7A", fill_type="solid")
_POS_FONT = Font(bold=True, color="375623")   # dark green
_NEG_FONT = Font(bold=True, color="9C0006")   # dark red


def calculate_net_totals(results: list[MatchResult]) -> dict[str, Decimal]:
    net: dict[str, Decimal] = defaultdict(Decimal)
    for r in results:
        wse = r.ws_record
        if wse and wse.rec_ccy and wse.rec_amount:
            net[wse.rec_ccy] += wse.rec_amount
        if wse and wse.pay_ccy and wse.pay_amount:
            net[wse.pay_ccy] -= wse.pay_amount
    return dict(net)


def format_net_figures(results: list[MatchResult]) -> str:
    net = calculate_net_totals(results)
    positives = {ccy: amt for ccy, amt in net.items() if amt >= 0}
    negatives = {ccy: amt for ccy, amt in net.items() if amt < 0}
    lines = []
    for ccy in sorted(positives):
        lines.append(f"{ccy:<6} {positives[ccy]:,.2f}")
    for ccy in sorted(negatives):
        lines.append(f"{ccy:<6} {negatives[ccy]:,.2f}")
    return "\n".join(lines)


def generate_payment_breakdown(
    results: list[MatchResult],
    output_path: str,
    report_date: date,
):
    """
    Payment breakdown sheet layout:

    Left:  flat table of all payments sorted alphabetically by buy currency.
           Columns: Conf# | Status | Value Date (GPG) | Buy Ccy | Buy Amt |
                    WS Value Date | Pay Ccy | Pay Amt | Rate | WS Deal # | WS Ext Deal #

    Right: net amounts grid starting two columns after the main table.
           Header: Ccy | Total
           One row per currency.
           Buy-side amounts (exotic received) → positive.
           Pay-side amounts (USD paid) → negative.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Breakdown"

    # ── Title block ───────────────────────────────────────────────────────────
    ws["A1"] = "Payment Breakdown Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Date:"
    ws["B2"] = report_date.strftime("%d %b %Y")
    ws["A3"] = "Total Records:"
    ws["B3"] = len(results)
    ws.append([])   # blank row 4

    # ── Main table headers (row 5) ────────────────────────────────────────────
    MAIN_HEADERS = [
        "Conf #", "Status",
        "Value Date (GPG)", "Buy Ccy", "Buy Amount",
        "Value Date (WS)", "Pay Ccy", "Pay Amount", "Rate",
        "WS Deal #", "WS Ext Deal #",
    ]
    HDR_ROW = 5
    for col_idx, h in enumerate(MAIN_HEADERS, 1):
        cell = ws.cell(row=HDR_ROW, column=col_idx, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Sort all results: alphabetically by buy_ccy, then by conf# ───────────
    def _sort_key(r: MatchResult):
        ccy = (
            r.gpg_record.buy_currency if r.gpg_record
            else r.ws_record.rec_ccy if r.ws_record
            else "ZZZ"
        )
        conf = (
            r.gpg_record.confirmation_number if r.gpg_record
            else _ws_display_key(r.ws_record) if r.ws_record
            else ""
        )
        return (ccy, conf)

    sorted_results = sorted(results, key=_sort_key)

    # ── Fill rows ─────────────────────────────────────────────────────────────
    net = calculate_net_totals(sorted_results)

    data_start_row = HDR_ROW + 1
    for r in sorted_results:
        gpg = r.gpg_record
        wse = r.ws_record
        conf = (
            gpg.confirmation_number if gpg
            else _ws_display_key(wse) if wse
            else ""
        )
        buy_ccy = (
            gpg.buy_currency if gpg
            else wse.rec_ccy if wse
            else ""
        )
        buy_amount = (
            gpg.buy_amount if gpg
            else wse.rec_amount if wse
            else Decimal("0")
        )

        row = [
            conf,
            r.status.value,
            gpg.value_date.strftime("%d %b %Y") if gpg else "",
            buy_ccy,
            float(buy_amount) if buy_amount else "",
            wse.value_date.strftime("%d %b %Y") if wse else "",
            wse.pay_ccy if wse else "",
            float(wse.pay_amount) if wse else "",
            float(wse.rate) if wse else "",
            wse.wallstreet_ref if wse else "",
            wse.external_ref if wse else "",
        ]
        ws.append(row)

    data_end_row = ws.max_row

    # Format amount columns (E=5, H=8) as numbers with 2 dp
    for row_idx in range(data_start_row, data_end_row + 1):
        for col_idx in (5, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.number_format = "#,##0.00"

    # ── Column widths for main table ──────────────────────────────────────────
    col_widths = [22, 18, 16, 10, 16, 16, 10, 16, 14, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Net grid (starts 2 columns after main table) ───────────────────────
    NET_START_COL = len(MAIN_HEADERS) + 2   # column N (14)
    net_hdr_row = HDR_ROW

    # Header
    for col_offset, label in enumerate(["Ccy", "Total"], 0):
        cell = ws.cell(row=net_hdr_row, column=NET_START_COL + col_offset, value=label)
        cell.fill = _NET_HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    # Rows: buy currencies (positive) first sorted, then pay currencies (negative)
    positives = {ccy: amt for ccy, amt in net.items() if amt >= 0}
    negatives = {ccy: amt for ccy, amt in net.items() if amt < 0}

    net_row = net_hdr_row + 1
    for ccy in sorted(positives):
        ws.cell(row=net_row, column=NET_START_COL, value=ccy).font = _TOTAL_FONT
        val_cell = ws.cell(row=net_row, column=NET_START_COL + 1, value=float(positives[ccy]))
        val_cell.number_format = "#,##0.00"
        val_cell.font = _POS_FONT
        net_row += 1

    for ccy in sorted(negatives):
        ws.cell(row=net_row, column=NET_START_COL, value=ccy).font = _TOTAL_FONT
        val_cell = ws.cell(row=net_row, column=NET_START_COL + 1, value=float(negatives[ccy]))
        val_cell.number_format = "#,##0.00"
        val_cell.font = _NEG_FONT
        net_row += 1

    # Net grid column widths
    ws.column_dimensions[get_column_letter(NET_START_COL)].width = 10
    ws.column_dimensions[get_column_letter(NET_START_COL + 1)].width = 18

    # Freeze panes at row 6 so headers stay visible
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    wb.save(output_path)


def _ws_display_key(wse) -> str:
    return wse.external_ref or wse.wallstreet_ref or ""
